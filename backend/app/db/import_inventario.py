"""Excel importer for the initial inventory migration (US 1.9).

Reads the legacy ``INVENTARIO_OFICIAL_OLIMPIA`` spreadsheet (or any workbook that
follows the same column layout) and creates ``empresa`` rows, writing an audit-log
INSERT snapshot for each — exactly like ``crud.create_empresa`` does.

Design notes:
- Column mapping is **header-driven** and accent/case-insensitive, so the operator
  does not have to rename columns to an exact casing.
- Category is resolved per row from a ``categoria`` column (slug *or* display name).
- Per-category extra columns are folded into ``campos_extras`` (mirrors the shape in
  ``frontend/components/empresa-form-fields.tsx``).
- **Idempotent**: a row whose ``nome_fantasia`` (accent/case-insensitive) already
  exists is skipped, so re-running the import never duplicates the trade.
- ``dry_run=True`` validates and reports without writing anything.

CLI:
    conda run -n oto python -m app.db.import_inventario caminho/arquivo.xlsx
    conda run -n oto python -m app.db.import_inventario arquivo.xlsx --dry-run
"""

from __future__ import annotations

import unicodedata
import uuid
from dataclasses import dataclass, field
from pathlib import Path

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.crud.audit import record_audit
from app.models.inventario import CategoriaEmpresa, Empresa

# ── Header normalisation ────────────────────────────────────────────────────


def _norm(text: object) -> str:
    """Lowercase, strip accents, collapse whitespace — for matching headers/values."""
    s = str(text or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.split())


# ── Column mapping ──────────────────────────────────────────────────────────

# Normalised header -> Empresa attribute. Multiple aliases may map to one field.
COLUMN_ALIASES: dict[str, str] = {
    "nome fantasia": "nome_fantasia",
    "nome": "nome_fantasia",
    "razao social": "razao_social",
    "cnpj": "cnpj",
    "endereco": "endereco",
    "bairro": "bairro",
    "telefone": "telefone",
    "email": "email",
    "e-mail": "email",
    "categoria": "_categoria",
    "aceita pesquisas": "aceita_pesquisas",
    "contato pesquisas": "contato_pesquisas",
    "contato para pesquisas": "contato_pesquisas",
    "telefone pesquisas": "telefone_pesquisas",
    "telefone para pesquisas": "telefone_pesquisas",
    "email pesquisas": "email_pesquisas",
    "email para pesquisas": "email_pesquisas",
    "proprietario": "proprietario",
}

# Per-category extra columns: normalised header -> (campos_extras key, kind).
# kind is "int" or "str"; mirrors CATEGORY_FIELDS in the frontend.
CATEGORY_EXTRA_COLUMNS: dict[str, dict[str, tuple[str, str]]] = {
    "meios_hospedagem": {
        "tipo": ("tipo", "str"),
        "tipo de hospedagem": ("tipo", "str"),
        "uhs": ("uhs", "int"),
        "unidades habitacionais": ("uhs", "int"),
        "leitos": ("leitos", "int"),
        "numero de leitos": ("leitos", "int"),
    },
    "alimentacao": {
        "capacidade": ("capacidade", "int"),
        "tipo de culinaria": ("tipo_culinaria", "str"),
        "culinaria": ("tipo_culinaria", "str"),
    },
    "atrativos": {
        "tipo": ("tipo", "str"),
        "tipo de atrativo": ("tipo", "str"),
    },
    "agencias": {
        "tipo": ("tipo", "str"),
    },
    "eventos": {
        "capacidade": ("capacidade_pessoas", "int"),
        "capacidade de pessoas": ("capacidade_pessoas", "int"),
    },
    "servicos_apoio": {
        "subcategoria": ("subcategoria", "str"),
    },
}

_TRUE_VALUES = {"sim", "s", "true", "verdadeiro", "1", "yes", "y"}
_FALSE_VALUES = {"nao", "n", "false", "falso", "0", "no"}


# ── Report types ────────────────────────────────────────────────────────────


@dataclass
class RowError:
    row: int  # 1-based row number in the sheet (header = row 1)
    message: str


@dataclass
class ImportReport:
    created: int = 0
    skipped: int = 0
    errors: list[RowError] = field(default_factory=list)

    @property
    def total(self) -> int:
        return self.created + self.skipped + len(self.errors)

    def summary(self) -> str:
        lines = [
            f"Importação concluída: {self.created} criada(s), "
            f"{self.skipped} já existente(s) ignorada(s), {len(self.errors)} com erro."
        ]
        for e in self.errors:
            lines.append(f"  linha {e.row}: {e.message}")
        return "\n".join(lines)


# ── Value coercion ──────────────────────────────────────────────────────────


def _clean_str(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_int(value: object) -> int | None:
    s = _clean_str(value)
    if s is None:
        return None
    # tolerate "240.0" coming from Excel numeric cells, and thousands separators
    s = s.replace(".", "").replace(",", "")
    return int(s)


def _parse_bool(value: object, default: bool = True) -> bool:
    s = _norm(value)
    if not s:
        return default
    if s in _TRUE_VALUES:
        return True
    if s in _FALSE_VALUES:
        return False
    return default


# ── Core import ─────────────────────────────────────────────────────────────


def _load_rows(path: Path) -> list[dict[str, object]]:
    """Read the first worksheet into a list of {normalised_header: value} dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = [_norm(h) for h in header]
    rows: list[dict[str, object]] = []
    for raw in rows_iter:
        rows.append({headers[i]: raw[i] for i in range(len(headers)) if headers[i]})
    wb.close()
    return rows


def _build_extras(slug: str, row: dict[str, object]) -> dict:
    extras: dict[str, object] = {}
    for header, (key, kind) in CATEGORY_EXTRA_COLUMNS.get(slug, {}).items():
        if header not in row:
            continue
        if kind == "int":
            parsed = _parse_int(row[header])
        else:
            parsed = _clean_str(row[header])
        if parsed is not None:
            extras[key] = parsed
    return extras


def import_workbook(
    db: Session,
    path: str | Path,
    *,
    usuario_id: uuid.UUID | None = None,
    dry_run: bool = False,
) -> ImportReport:
    """Import an inventory workbook into ``empresa``. See module docstring."""
    path = Path(path)
    report = ImportReport()

    categorias = {c.slug: c for c in db.query(CategoriaEmpresa).all()}
    categorias_by_nome = {_norm(c.nome): c for c in categorias.values()}

    rows = _load_rows(path)
    # track names seen within this run too, so duplicates inside the file are caught
    seen_in_run: set[str] = set()

    for idx, row in enumerate(rows, start=2):  # +2: header is row 1, data starts at 2
        # fully empty row → silently skip
        if not any(_clean_str(v) for v in row.values()):
            continue

        nome = _clean_str(row.get("nome fantasia") or row.get("nome"))
        if not nome:
            report.errors.append(RowError(idx, "nome_fantasia ausente"))
            continue

        cat_raw = _clean_str(row.get("categoria"))
        if not cat_raw:
            report.errors.append(RowError(idx, "categoria ausente"))
            continue
        cat_norm = _norm(cat_raw)
        categoria = categorias.get(cat_norm) or categorias_by_nome.get(cat_norm)
        if categoria is None:
            report.errors.append(RowError(idx, f"categoria desconhecida: {cat_raw!r}"))
            continue

        nome_key = _norm(nome)
        if nome_key in seen_in_run:
            report.skipped += 1
            continue
        existing = (
            db.query(Empresa.id)
            .filter(func.unaccent(func.lower(Empresa.nome_fantasia)) == func.unaccent(nome_key))
            .first()
        )
        if existing is not None:
            report.skipped += 1
            seen_in_run.add(nome_key)
            continue

        empresa = Empresa(
            categoria_id=categoria.id,
            nome_fantasia=nome,
            razao_social=_clean_str(row.get("razao social")),
            cnpj=_clean_str(row.get("cnpj")),
            endereco=_clean_str(row.get("endereco")),
            bairro=_clean_str(row.get("bairro")),
            telefone=_clean_str(row.get("telefone")),
            email=_clean_str(row.get("email") or row.get("e-mail")),
            aceita_pesquisas=_parse_bool(row.get("aceita pesquisas")),
            contato_pesquisas=_clean_str(
                row.get("contato pesquisas") or row.get("contato para pesquisas")
            ),
            telefone_pesquisas=_clean_str(
                row.get("telefone pesquisas") or row.get("telefone para pesquisas")
            ),
            email_pesquisas=_clean_str(
                row.get("email pesquisas") or row.get("email para pesquisas")
            ),
            proprietario=_clean_str(row.get("proprietario")),
            campos_extras=_build_extras(categoria.slug, row) or None,
            criado_por=usuario_id,
        )
        db.add(empresa)
        db.flush()  # populate empresa.id before audit

        snapshot = {
            "categoria_id": categoria.id,
            "nome_fantasia": empresa.nome_fantasia,
            "razao_social": empresa.razao_social,
            "cnpj": empresa.cnpj,
            "campos_extras": empresa.campos_extras,
            "origem": "import_excel",
        }
        record_audit(db, "empresa", empresa.id, usuario_id, "INSERT", valor_novo=snapshot)

        report.created += 1
        seen_in_run.add(nome_key)

    if dry_run:
        db.rollback()
    else:
        db.commit()
    return report


def write_template(dest: str | Path) -> Path:
    """Write a blank workbook with the column headers the importer understands.

    Includes the common empresa columns plus one example of each category's extra
    columns, so the operator can fill it in and re-import.
    """
    from openpyxl import Workbook

    headers = [
        "Nome Fantasia", "Categoria", "Razao Social", "CNPJ", "Endereco", "Bairro",
        "Telefone", "Email", "Proprietario", "Aceita Pesquisas",
        "Contato Pesquisas", "Telefone Pesquisas", "Email Pesquisas",
        # category-specific (only fill the ones relevant to the row's categoria):
        "Tipo", "UHs", "Leitos", "Capacidade", "Tipo de Culinaria", "Subcategoria",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "inventario"
    ws.append(headers)
    # a hint row showing valid category slugs
    ws.append(["Ex.: Hotel Modelo", "meios_hospedagem", "", "", "", "", "", "", "",
               "Sim", "", "", "", "hotel", 100, 200, "", "", ""])
    dest = Path(dest)
    wb.save(dest)
    return dest


def _main() -> None:
    import sys

    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    dry_run = "--dry-run" in sys.argv

    if "--template" in sys.argv:
        out = write_template(args[0] if args else "modelo_inventario.xlsx")
        print(f"Modelo gravado em {out}")
        return

    if not args:
        print("uso: python -m app.db.import_inventario <arquivo.xlsx> [--dry-run]")
        print("     python -m app.db.import_inventario [modelo.xlsx] --template")
        raise SystemExit(2)

    from app.db.seed import run_seed  # ensure categorias exist
    from app.db.session import SessionLocal
    from app.models.usuario import Usuario

    session = SessionLocal()
    try:
        run_seed(session)
        admin = session.query(Usuario).filter_by(perfil="admin").first()
        report = import_workbook(
            session, args[0], usuario_id=admin.id if admin else None, dry_run=dry_run
        )
        print(("[dry-run] " if dry_run else "") + report.summary())
    finally:
        session.close()


if __name__ == "__main__":
    _main()
